import pygame, sys, random
import box_control
from pygame.locals import *
import func
import time
from smalldata import *
import openpyxl
"""
import openpyxl
workbook=openpyxl.load_workbook("C:\\Users\\Arc en ciel\\Desktop\\icm\\orihinal\\newM.xlsx")
booksheet=workbook.get_sheet_by_name('Sheet1')
X = []
for row in booksheet.rows:
    tmp=[]
for i in range(len(row)):
    if(row[i].value!=None):
        tmp.append(row[i].value)
    X.append(tmp)
"""
#workbook=openpyxl.load_workbook("E:\\final\\X.xlsx")
#booksheet=workbook.get_sheet_by_name('Sheet1')
class person:
    id = -1
    x =-1
    y = -1
    xe = -1
    ye = -1
    fspeed=-1
    delay=-1
    steps=0
    def __init__(self, x1, y1):
        self.id = -1
        self.x =-1
        self.y = -1
        self.xe = -1
        self.ye = -1
        self.fspeed=-1
        self.delay=0
    def setXY(self,x1,y1,id):
        self.id=id
        self.x=x1
        self.y=y1
    def stXY(self,x1,y1):
        self.x=x1
        self.y=y1
    def getX(self):
        return self.x
    def getY(self):
        return self.y
    def setfspeed(self,spe):
        self.fspeed=spe
    def setde(self,dela):
        self.delay=dela
    def add(self):
        self.steps+=1
def flush_box(box_color, rect):  #涂色
    pygame.draw.rect(win, box_color, rect)
win_width = 532*2
win_height = 216*2
#win_width = 668+3+3wor
#win_height = 268+3+wor
box_numx = 64 #网wor
box_numy = 51 #网格
life_num =200#人数
change_start = True
pygame.init()
mod=1
id=0
boxs,people_loc= box_control.init_boxs(box_numx,box_numy, win_width, win_height, life_num)
win = pygame.display.set_mode((win_width, win_height))
pygame.display.set_caption('zhan')
clock = pygame.time.Clock()
people = []
for i in range (0,life_num):
    tmp = person(-1, -1)
    people.append(tmp)
print(len(people))
peo={}
for i in range(len(people)):
    people[i].setXY(people_loc[i][0],people_loc[i][1],i)
    print(people[i].x,people[i].y)
    if(i%mod==0):
        people[i].setfspeed(1)
        people[i].setde(0)
        boxs[people[i].x][people[i].y]['value']=1
    elif(i%mod==1):
        people[i].setfspeed(1)
        people[i].setde(2)
        boxs[people[i].x][people[i].y]['value']=11
    else:
        people[i].setfspeed(0.5)
        people[i].setde(0)
        boxs[people[i].x][people[i].y]['value']=0.1
dct={}
for x in range(0,box_numx):
    for y in range(0,box_numy):
        dct[boxs[x][y]['value']]=0
for x in range(0,box_numx):
    for y in range(0,box_numy):
        dct[boxs[x][y]['value']]+=1
print("Number of People:",dct[1])
#booksheet.cell(1,10,"Number")
#booksheet.cell(1,11,dct[1])

print(enter)
cc = 0
numofper=0
while True:
    clock.tick(1)
    persons=0
    for event in pygame.event.get():
        if event.type == QUIT:
            pygame.quit()
            sys.exit()
        if event.type == KEYDOWN and event.key == K_SPACE:
            change_start = not change_start
    dis=0
    if change_start:
        f=cc%2
        for x in range(0, box_numx):
            for y in range(0, box_numy):
                if boxs[x][y]['value'] == -1:
                    color = (0,0,0) #
                    
                elif boxs[x][y]['value'] == 1:
                    color = (25,25,112)  
                    dis=1
                elif boxs[x][y]['value'] == 0.1:
                    color = (220,20,60)  
                    dis=1
                elif boxs[x][y]['value'] == 11:
                    color = (173,255,47)  
                    dis=1
                elif boxs[x][y]['value'] == 2:
                    color = (255,130,71)  
                    
                else:
                    color = (255,255,255) 
                flush_box(color,
                          pygame.Rect(
                              boxs[x][y]['x'],
                              boxs[x][y]['y'],
                              boxs[x][y]['width'],
                              boxs[x][y]['height'],
                          ))
        box_control.change_boxs(boxs,people,enter,f,cc)
        if(not dis):
            print("TotalTime",cc)  
            #booksheet.cell(2,10,"TotalTime")
            #booksheet.cell(2,11,cc)
            break
            #for i in range(len(people)):
                #booksheet.cell(i+3,10,i)
                #booksheet.cell(i+3,11,people[i].steps)
            #workbook.save("E:\\final\\X.xlsx")
        cc += 1
    pygame.display.update()

